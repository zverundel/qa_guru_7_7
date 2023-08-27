from openpyxl import load_workbook
from .conftest import RESOURCES_DIR
import os.path


# TODO оформить в тест, добавить ассерты и использовать универсальный путь


def test_xlsx_file():
    workbook = load_workbook(os.path.join(RESOURCES_DIR, 'file_example_XLSX_50.xlsx'))
    sheet = workbook.active
    print(sheet.cell(row=3, column=2).value)
    assert sheet.cell(row=3, column=2).value == 'Mara'
